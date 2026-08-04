"""Import the org's member roster (the student database xlsx) into the portal.

Each named row with a Gmail becomes a login account (User) plus a
MemberProfile holding the biographical/contact fields. Rows with no Gmail are
skipped — email is the portal's login identity. Because the roster is the
org's own verified list, imported accounts are pre-linked to Google
(google_linked_at set) so members can sign straight in with their Gmail, and
are given the "Member" access level.

Idempotent: matches existing accounts by email and updates them, so re-running
never creates duplicates.

Usage (from backend/, with the venv):
    DATABASE_URL=sqlite:///./portal_dev.db \\
        .venv/bin/python -m scripts.import_members "/path/to/Members database.xlsx"
    # add --commit to write; without it, does a dry run and reports counts.

The sheet layout is fixed: column A = MTR id, and row 1 columns B..N are the
field headers (Name, National ID, Birthday, ...). The reserved "gap" rows
(no Name) are ignored automatically.
"""

import argparse
import secrets
import sys
from datetime import date, datetime, timezone

import openpyxl
from sqlalchemy import select

from app.core.database import SessionLocal
from app.core.security import hash_password
from app.domains.access.models import AccessLevel
from app.domains.users.models import MemberProfile, User

# roster header text -> MemberProfile attribute. Column A is always the MTR id
# regardless of what's in the A1 cell (which holds an id, not a header).
HEADER_TO_FIELD = {
    "national id": "national_id",
    "birthday": "birthday",
    "university": "university",
    "college": "college",
    "major": "major",
    "graduating year": "graduating_year",
    "phone number": "phone",
    "dad's number": "father_phone",
    "mom's number": "mother_phone",
    "uni id": "uni_id",
    "location": "location",
}
MEMBER_LEVEL_NAME = "Member"


def _text(value) -> str | None:
    """Coerce a cell to a trimmed string, dropping the trailing .0 Excel adds
    to whole numbers (grad years, UNI/national IDs read as floats). Phone
    numbers keep their leading zeros because Excel already stores them as text."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _year(value) -> int | None:
    text = _text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _birthday(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _column_map(ws) -> dict[int, str]:
    """Map 1-based column index -> profile field, by reading row-1 headers."""
    mapping: dict[int, str] = {}
    for col in range(2, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header is None:
            continue
        key = str(header).strip().lower()
        if key == "name":
            mapping[col] = "__name__"
        elif key == "gmail":
            mapping[col] = "__email__"
        elif key in HEADER_TO_FIELD:
            mapping[col] = HEADER_TO_FIELD[key]
    return mapping


def run(path: str, commit: bool) -> int:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    colmap = _column_map(ws)

    db = SessionLocal()
    member_level = db.scalar(select(AccessLevel).where(AccessLevel.name == MEMBER_LEVEL_NAME))
    if member_level is None:
        print(f"ERROR: no '{MEMBER_LEVEL_NAME}' access level found — is the DB seeded?")
        return 1

    created = updated = skipped_no_email = skipped_no_name = 0
    seen_emails: set[str] = set()
    duplicate_emails: list[str] = []
    errors: list[str] = []

    for row in range(2, ws.max_row + 1):
        mtr_id = _text(ws.cell(row=row, column=1).value)
        raw = {field: ws.cell(row=row, column=col).value for col, field in colmap.items()}

        name = _text(raw.get("__name__"))
        if not name:
            skipped_no_name += 1
            continue  # reserved gap / blank row

        email = _text(raw.get("__email__"))
        if not email:
            skipped_no_email += 1
            continue  # no login identity — skipped per import policy
        email = email.lower()
        if email in seen_emails:
            duplicate_emails.append(f"{email} (row {row}, {name})")
        seen_emails.add(email)

        fields = {
            "mtr_id": mtr_id,
            "national_id": _text(raw.get("national_id")),
            "birthday": _birthday(raw.get("birthday")),
            "university": _text(raw.get("university")),
            "college": _text(raw.get("college")),
            "major": _text(raw.get("major")),
            "graduating_year": _year(raw.get("graduating_year")),
            "phone": _text(raw.get("phone")),
            "father_phone": _text(raw.get("father_phone")),
            "mother_phone": _text(raw.get("mother_phone")),
            "uni_id": _text(raw.get("uni_id")),
            "location": _text(raw.get("location")),
        }

        user = db.scalar(select(User).where(User.email == email))
        if user is None:
            user = User(
                email=email,
                full_name=name,
                hashed_password=hash_password(secrets.token_urlsafe(32)),
                access_level_id=member_level.id,
                google_linked_at=datetime.now(timezone.utc),
                is_active=True,
            )
            db.add(user)
            db.flush()
            created += 1
        else:
            updated += 1

        profile = user.profile
        if profile is None:
            profile = MemberProfile()
            user.profile = profile  # keeps the 1:1 consistent within this run
        for field, value in fields.items():
            setattr(profile, field, value)

    try:
        db.flush()
    except Exception as exc:  # surface a bad row rather than a silent rollback
        errors.append(repr(exc))

    print("Import summary")
    print(f"  accounts created : {created}")
    print(f"  accounts updated : {updated}")
    print(f"  skipped (no Gmail): {skipped_no_email}")
    print(f"  skipped (no name / gap rows): {skipped_no_name}")
    if duplicate_emails:
        print(f"  duplicate Gmails in sheet (last row wins): {len(duplicate_emails)}")
        for d in duplicate_emails:
            print(f"    - {d}")
    if errors:
        print("  errors:")
        for e in errors:
            print(f"    - {e}")

    if commit and not errors:
        db.commit()
        print("COMMITTED.")
    else:
        db.rollback()
        print("DRY RUN — nothing written." if not commit else "ERRORS — rolled back.")
    db.close()
    return 1 if errors else 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--commit", action="store_true", help="write to the DB (default: dry run)")
    args = ap.parse_args()
    sys.exit(run(args.path, args.commit))
