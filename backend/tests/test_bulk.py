"""Bulk-editor engine: read / validate / apply, privilege gating, and the
per-table safety rules (append-only ledger, delete guards, deactivate-not-
delete for people)."""

from sqlalchemy import select

from app.domains.inventory.models import InventoryItem, StockMovement
from app.domains.users.models import User


# --- listing / privilege gating --------------------------------------------
def test_list_tables_filtered_by_privilege(login):
    exec_tables = {t["key"] for t in login("ceo").get("/api/bulk/tables").json()}
    # Exec holds inventory.edit, org.edit, events.manage_any, users.manage? No —
    # users.manage is admin-only in the ladder, so 'people' is hidden for Exec.
    assert "inventory_items" in exec_tables
    assert "events" in exec_tables
    assert "people" not in exec_tables  # needs users.manage

    admin_tables = {t["key"] for t in login("admin").get("/api/bulk/tables").json()}
    assert "people" in admin_tables  # admin holds everything

    member_tables = login("comp_member").get("/api/bulk/tables").json()
    assert member_tables == []  # no edit privileges -> nothing


def test_get_table_requires_privilege(login):
    # Member can't read the inventory_items bulk grid (no inventory.edit)
    assert login("comp_member").get("/api/bulk/inventory_items").status_code == 403
    assert login("ceo").get("/api/bulk/inventory_items").status_code == 200


def test_unknown_table_404(login):
    assert login("admin").get("/api/bulk/nope").status_code == 404


# --- Google Sheets escape hatch ---------------------------------------------
def _force_sheets_unconfigured(monkeypatch):
    # Don't depend on the ambient backend/.env (a real dev box may have Sheets
    # configured) — pin the unconfigured state for these tests.
    from app.core.config import settings

    monkeypatch.setattr(settings, "google_sheets_credentials_file", "")
    monkeypatch.setattr(settings, "google_sheets_credentials_b64", "")
    monkeypatch.setattr(settings, "google_sheets_spreadsheet_id", "")


def test_sheets_status_reports_unconfigured(login, monkeypatch):
    # No creds / spreadsheet id -> not configured, so the UI shows the Sheets
    # buttons disabled instead of erroring.
    _force_sheets_unconfigured(monkeypatch)
    body = login("admin").get("/api/bulk/sheets/status").json()
    assert body["configured"] is False


def test_sheet_push_400_when_unconfigured(login, monkeypatch):
    _force_sheets_unconfigured(monkeypatch)
    r = login("ceo").post("/api/bulk/inventory_items/sheet/push")
    assert r.status_code == 400
    assert "configured" in r.json()["detail"].lower()


def test_sheet_pull_requires_privilege(login):
    # gated by the same per-table privilege as the rest of the bulk editor
    assert login("comp_member").post("/api/bulk/inventory_items/sheet/pull").status_code == 403


# --- live sheet -> DB webhook -----------------------------------------------
def test_sheet_webhook_disabled_without_token(login, monkeypatch):
    from app.core.config import settings

    monkeypatch.setattr(settings, "sheets_sync_token", "")  # webhook disabled -> 404
    r = login("admin").post(
        "/api/bulk/sheet-webhook", json={"tab": "inventory_items", "row": {}}
    )
    assert r.status_code == 404


def test_sheet_webhook_rejects_bad_token(login, monkeypatch):
    from app.core.config import settings

    monkeypatch.setattr(settings, "sheets_sync_token", "s3cret")
    r = login("admin").post(
        "/api/bulk/sheet-webhook",
        json={"tab": "inventory_items", "row": {"id": "", "name": "X", "quantity": "1"}},
        headers={"X-Sheet-Token": "wrong"},
    )
    assert r.status_code == 401


def test_sheet_webhook_inserts_then_updates(login, db_session, monkeypatch):
    from app.core.config import settings
    from app.domains.inventory.models import InventoryItem

    monkeypatch.setattr(settings, "sheets_sync_token", "s3cret")
    admin = login("admin")
    hdr = {"X-Sheet-Token": "s3cret"}

    # a blank-id edit inserts and returns the new id (script writes it back)
    r = admin.post(
        "/api/bulk/sheet-webhook",
        json={"tab": "inventory_items", "row": {"id": "", "name": "Live Widget", "quantity": "4"}},
        headers=hdr,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True and isinstance(body["id"], int)
    new_id = body["id"]
    db_session.expire_all()
    assert db_session.get(InventoryItem, new_id).quantity == 4

    # editing that same row (now carrying the id) updates in place, no duplicate
    r = admin.post(
        "/api/bulk/sheet-webhook",
        json={
            "tab": "inventory_items",
            "row": {"id": str(new_id), "name": "Live Widget", "quantity": "9"},
        },
        headers=hdr,
    )
    assert r.json()["id"] == new_id
    db_session.expire_all()
    assert db_session.get(InventoryItem, new_id).quantity == 9


def test_sheet_webhook_reports_validation_error(login, monkeypatch):
    from app.core.config import settings

    monkeypatch.setattr(settings, "sheets_sync_token", "s3cret")
    r = login("admin").post(
        "/api/bulk/sheet-webhook",
        json={"tab": "inventory_items", "row": {"id": "", "name": "", "quantity": "x"}},
        headers={"X-Sheet-Token": "s3cret"},
    )
    body = r.json()
    assert body["ok"] is False and body["id"] is None
    assert any(e["column"] in ("name", "quantity") for e in body["errors"])


# --- validate ---------------------------------------------------------------
def test_validate_catches_required_and_type_and_bad_ref(login):
    r = login("ceo").post("/api/bulk/inventory_items/validate", json={
        "rows": [
            {"id": "", "name": "", "quantity": "x"},                       # missing name, bad int
            {"id": "", "name": "Good", "quantity": "3", "team_lead_id": "999999"},  # bad ref
        ],
        "deletes": [],
    })
    assert r.status_code == 200
    errors = r.json()["errors"]
    cols = {(e["row"], e["column"]) for e in errors}
    assert (0, "name") in cols
    assert (0, "quantity") in cols
    assert (1, "team_lead_id") in cols
    assert r.json()["ok"] is False


def test_validate_bad_choice(login):
    r = login("ceo").post("/api/bulk/inventory_items/validate", json={
        "rows": [{"id": "", "name": "X", "quantity": "1", "condition": "melted"}],
    })
    assert any(e["column"] == "condition" for e in r.json()["errors"])


# --- apply: add / update ----------------------------------------------------
def test_apply_add_and_update(login, db_session):
    before = db_session.scalar(select(InventoryItem.id).order_by(InventoryItem.id.desc()))
    ceo = login("ceo")
    r = ceo.post("/api/bulk/inventory_items", json={
        "rows": [
            {"id": "", "name": "Bulk Widget", "quantity": "7", "condition": "good"},
        ],
    })
    assert r.status_code == 200, r.text
    assert r.json()["ok"] is True
    assert r.json()["summary"]["adds"] == 1

    item = db_session.scalar(select(InventoryItem).where(InventoryItem.name == "Bulk Widget"))
    assert item is not None and item.quantity == 7

    # update the same row by id
    r = ceo.post("/api/bulk/inventory_items", json={
        "rows": [{"id": str(item.id), "name": "Bulk Widget", "quantity": "12", "condition": "fair"}],
    })
    assert r.json()["summary"]["updates"] == 1
    db_session.expire_all()
    refreshed = db_session.get(InventoryItem, item.id)
    assert refreshed.quantity == 12 and refreshed.condition == "fair"


def test_apply_rejects_wholesale_on_any_error(login, db_session):
    n_before = db_session.scalar(select(InventoryItem.id).order_by(InventoryItem.id.desc()))
    r = login("ceo").post("/api/bulk/inventory_items", json={
        "rows": [
            {"id": "", "name": "Would Be Added", "quantity": "1"},
            {"id": "", "name": "", "quantity": "1"},  # invalid -> whole save aborts
        ],
    })
    assert r.json()["ok"] is False
    # nothing written
    assert db_session.scalar(select(InventoryItem).where(InventoryItem.name == "Would Be Added")) is None


# --- people: deactivate, never hard delete ----------------------------------
def test_delete_people_deactivates(login, db_session, org):
    target = org["comp_member"].id
    r = login("admin").post("/api/bulk/people", json={"rows": [], "deletes": [target]})
    assert r.json()["ok"] is True and r.json()["summary"]["deletes"] == 1
    db_session.expire_all()
    user = db_session.get(User, target)
    assert user is not None  # still present
    assert user.is_active is False  # just deactivated


# --- movements: append-only ledger ------------------------------------------
def test_movements_append_only(login, db_session, org):
    item = InventoryItem(name="Ledger Item", quantity=10)
    db_session.add(item)
    db_session.commit()
    ceo = login("ceo")
    # insert a new ledger row (stock-in): from nowhere -> nowhere, qty 5
    r = ceo.post("/api/bulk/inventory_movements", json={
        "rows": [{"id": "", "item_id": str(item.id), "quantity": "5", "reason": "bulk intake"}],
    })
    assert r.json()["ok"] is True, r.text
    assert r.json()["summary"]["adds"] == 1
    mv = db_session.scalar(select(StockMovement).where(StockMovement.reason == "bulk intake"))
    assert mv is not None and mv.quantity == 5

    # deleting a ledger row is refused
    r = ceo.post("/api/bulk/inventory_movements", json={"rows": [], "deletes": [mv.id]})
    assert r.json()["ok"] is False
    assert any("append-only" in e["message"] for e in r.json()["errors"])


# --- delete guards ----------------------------------------------------------
def test_delete_guard_blocks_referenced_kind(login, db_session):
    # Event/Training/R&D kinds exist from ensure_preset_kinds; create an event
    # under one, then try to delete the kind -> blocked.
    from app.domains.events.models import Event, EventKind
    kind = db_session.scalar(select(EventKind))
    db_session.add(Event(name="Guarded Ev", kind_id=kind.id))
    db_session.commit()
    r = login("ceo").post("/api/bulk/event_kinds", json={"rows": [], "deletes": [kind.id]})
    assert r.json()["ok"] is False
    assert any("events still use it" in e["message"] for e in r.json()["errors"])


# --- Excel round-trip -------------------------------------------------------
def _read_xlsx(content: bytes):
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    grid = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    return ws.title, grid


def _write_xlsx(header, rows) -> bytes:
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_export_xlsx_has_header_and_rows(login, db_session):
    from app.domains.inventory.models import InventoryItem

    db_session.add(InventoryItem(name="Exported Widget", quantity=3))
    db_session.commit()
    r = login("ceo").get("/api/bulk/inventory_items/export.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    title, grid = _read_xlsx(r.content)
    assert title == "inventory_items"
    assert grid[0][:2] == ["id", "name"]
    assert any(row[1] == "Exported Widget" for row in grid[1:])


def test_upload_preview_then_apply(login, db_session):
    from sqlalchemy import select

    from app.domains.inventory.models import InventoryItem

    db_session.add(InventoryItem(name="Seed A", quantity=1))
    db_session.commit()
    ceo = login("ceo")

    header = ["id", "name", "category", "asset_tag", "sku", "quantity", "low_stock_threshold",
              "unit", "location", "condition", "notes", "team_lead_id"]
    seed = db_session.scalar(select(InventoryItem).where(InventoryItem.name == "Seed A"))
    rows = [
        [seed.id, "Seed A EDITED", "", "", "", 9, 0, "unit", "", "good", "", ""],  # update
        ["", "Fresh From Excel", "", "", "", 5, 0, "unit", "", "good", "", ""],     # add
    ]
    content = _write_xlsx(header, rows)

    # preview only — no writes
    files = {"file": ("inventory_items.xlsx", content, "application/octet-stream")}
    prev = ceo.post("/api/bulk/inventory_items/upload", files=files).json()
    assert prev["applied"] is False
    assert prev["new"] == 1 and prev["changed"] == 1
    db_session.expire_all()
    assert db_session.scalar(select(InventoryItem).where(InventoryItem.name == "Fresh From Excel")) is None

    # apply
    files = {"file": ("inventory_items.xlsx", content, "application/octet-stream")}
    res = ceo.post("/api/bulk/inventory_items/upload?apply=true", files=files).json()
    assert res["applied"] is True
    assert res["summary"]["adds"] == 1 and res["summary"]["updates"] == 1
    db_session.expire_all()
    assert db_session.scalar(select(InventoryItem).where(InventoryItem.name == "Fresh From Excel")) is not None
    assert db_session.get(InventoryItem, seed.id).quantity == 9


def test_upload_delete_missing_is_opt_in(login, db_session):
    from sqlalchemy import select

    from app.domains.inventory.models import InventoryItem

    keep = InventoryItem(name="Keep Me", quantity=1)
    drop = InventoryItem(name="Drop Me", quantity=1)
    db_session.add_all([keep, drop])
    db_session.commit()
    ceo = login("ceo")

    header = ["id", "name", "category", "asset_tag", "sku", "quantity", "low_stock_threshold",
              "unit", "location", "condition", "notes", "team_lead_id"]
    # upload a sheet with only "Keep Me" — "Drop Me" is absent
    content = _write_xlsx(header, [[keep.id, "Keep Me", "", "", "", 1, 0, "unit", "", "good", "", ""]])

    # preview reports the missing row but delete is off by default
    files = {"file": ("inventory_items.xlsx", content, "application/octet-stream")}
    prev = ceo.post("/api/bulk/inventory_items/upload", files=files).json()
    assert drop.id in prev["missing_ids"] and prev["will_delete"] == 0

    # apply WITHOUT delete_missing -> Drop Me survives
    files = {"file": ("inventory_items.xlsx", content, "application/octet-stream")}
    ceo.post("/api/bulk/inventory_items/upload?apply=true", files=files)
    db_session.expire_all()
    assert db_session.get(InventoryItem, drop.id).deleted_at is None

    # apply WITH delete_missing -> Drop Me is soft-deleted
    files = {"file": ("inventory_items.xlsx", content, "application/octet-stream")}
    res = ceo.post(
        "/api/bulk/inventory_items/upload?apply=true&delete_missing=true", files=files
    ).json()
    assert res["summary"]["deletes"] == 1
    db_session.expire_all()
    assert db_session.get(InventoryItem, drop.id).deleted_at is not None


def test_upload_missing_column_rejected(login, db_session):
    from app.domains.inventory.models import InventoryItem

    db_session.add(InventoryItem(name="X", quantity=1))
    db_session.commit()
    # header without 'quantity'
    content = _write_xlsx(["id", "name"], [["", "Nope"]])
    files = {"file": ("inventory_items.xlsx", content, "application/octet-stream")}
    r = login("ceo").post("/api/bulk/inventory_items/upload", files=files)
    assert r.status_code == 400
    assert "quantity" in r.json()["detail"]
