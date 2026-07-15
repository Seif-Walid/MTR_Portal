"""Parse an uploaded .xlsx/.csv into the same (headers, rows) shape the
Google Sheet importer produces, so both sources feed the same column-mapping
UI and the same import logic. No Google credentials required."""

import csv
import io

import openpyxl

Grid = list[list[str]]


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _grid_to_headers_and_rows(grid: Grid) -> tuple[list[str], list[dict[str, str]]]:
    header_idx = next(
        (i for i, row in enumerate(grid[:25]) if sum(1 for c in row if c.strip()) >= 2),
        None,
    )
    if header_idx is None:
        return [], []
    header_row = grid[header_idx]
    body = [row for row in grid[header_idx + 1 :] if any(c.strip() for c in row)]

    keep = [
        i
        for i in range(len(header_row))
        if header_row[i].strip() or any(i < len(r) and r[i].strip() for r in body)
    ]
    headers: list[str] = []
    seen: set[str] = set()
    for i in keep:
        label = header_row[i].strip() or f"Column {i + 1}"
        base, n = label, 2
        while label in seen:
            label = f"{base} ({n})"
            n += 1
        seen.add(label)
        headers.append(label)

    rows = [
        {headers[j]: (r[i] if i < len(r) else "") for j, i in enumerate(keep)}
        for r in body
    ]
    return headers, rows


def list_sheets(content: bytes, filename: str) -> list[str] | None:
    """Sheet names for a workbook, or None for single-sheet formats like CSV."""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return None
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    return wb.sheetnames


def read_uploaded(
    content: bytes, filename: str, sheet: str | None = None
) -> tuple[str | None, list[str], list[dict[str, str]]]:
    """Returns (chosen_sheet_name, headers, rows). Auto-detects the header row
    as the first one with at least two non-empty cells, and drops columns that
    are empty in both the header and every data row."""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        grid = [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
        headers, rows = _grid_to_headers_and_rows(grid)
        return ws.title, headers, rows
    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        grid = list(csv.reader(io.StringIO(text)))
        headers, rows = _grid_to_headers_and_rows(grid)
        return None, headers, rows
    raise ValueError("Unsupported file type — upload a .xlsx or .csv file.")
