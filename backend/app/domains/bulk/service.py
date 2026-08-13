"""Generic bulk-editor engine: read / validate / apply for any table in the
registry.

Unlike the Sheets *rebuild* (which truncates and recreates every managed table
from a full snapshot — see sync.service), this is an **incremental, id-keyed
upsert**: rows carrying an id are updated, rows without one are inserted, and
ids listed in `deletes` are removed per the table's delete policy. A save is
all-or-nothing — if validation finds any problem, nothing is written.
"""

from __future__ import annotations

import secrets
from dataclasses import dataclass
from datetime import date, datetime, timezone

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.security import hash_password
from app.domains.access.models import AccessLevel
from app.domains.audit import service as audit
from app.domains.events.models import Event, EventCategory, EventTeam, EventTeamMember
from app.domains.inventory.models import InventoryItem, StockMovement
from app.domains.positions.models import Position, PositionOccupant
from app.domains.users.models import User
from app.domains.sync import service as sync
from app.domains.sync.service import _EXPORTERS, _bool, _parse_date, _parse_datetime
from app.domains.bulk.registry import REFS, TABLES, Column, TableSpec


# --- reads ------------------------------------------------------------------
# Read-only auto-registered tables are capped so a huge table (audit log,
# notifications) can't blow up one response.
_READONLY_ROW_CAP = 2000


def _cell(value: object) -> str:
    """Stringify any DB value into the grid's string-cell convention."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _generic_export(db: Session, spec: TableSpec) -> tuple[list[str], list[list[str]]]:
    """Reflect any mapped table into (header, rows) without a hand-written
    exporter — used for the read-only 'every table' specs."""
    model = spec.model
    cols = [c.key for c in model.__table__.columns]
    objs = db.scalars(select(model).limit(_READONLY_ROW_CAP))
    rows = [[_cell(getattr(o, c)) for c in cols] for o in objs]
    return cols, rows


def _export_for(db: Session, spec: TableSpec) -> tuple[list[str], list[list[str]]]:
    exporter = _EXPORTERS.get(spec.key)
    return exporter(db) if exporter else _generic_export(db, spec)


def read_table(db: Session, spec: TableSpec) -> dict:
    """Current DB rows for a table, plus the dropdown options every reference
    column needs. Row serialization is the exact same code the Sheets export
    uses, so grid / CSV / mirror never disagree."""
    header, rows = _export_for(db, spec)
    columns = [_column_payload(c) for c in spec.columns]
    row_dicts = [dict(zip(header, r)) for r in rows]
    return {
        "key": spec.key,
        "label": spec.label,
        "columns": columns,
        "rows": row_dicts,
        "options": _ref_options(db, spec),
        "append_only": spec.append_only,
        "delete": spec.delete,
        "read_only": spec.read_only,
    }


def _column_payload(c: Column) -> dict:
    return {
        "name": c.name,
        "type": c.type,
        "editable": c.editable,
        "required": c.required,
        "ref": c.ref,
        "choices": list(c.choices) if c.choices else None,
    }


def _ref_options(db: Session, spec: TableSpec) -> dict[str, list[dict]]:
    """{column_name: [{value, label}, ...]} for every reference column, so the
    grid can render searchable dropdowns and match ids to human labels."""
    out: dict[str, list[dict]] = {}
    for col in spec.columns:
        if not col.ref or col.ref in out:
            continue
        ref = REFS[col.ref]
        opts = [
            {"value": getattr(o, ref.value_attr), "label": ref.label(o)}
            for o in db.scalars(select(ref.model))
        ]
        out[col.name] = opts
    return out


# --- validation -------------------------------------------------------------
@dataclass
class Coerced:
    """One validated incoming row: model attribute -> python value, ready to
    write. `index` is the row's position in the submitted array (for anchoring
    errors back to the grid); `id` is the existing row id, or None for inserts."""

    index: int
    id: int | None
    fields: dict


def _err(errors: list[dict], index: int | None, column: str | None, message: str) -> None:
    errors.append({"row": index, "column": column, "message": message})


def _valid_ref_ids(db: Session, ref_key: str, deleted: dict[str, set[int]]) -> set[int]:
    ref = REFS[ref_key]
    ids = {r[0] for r in db.execute(select(ref.model.id)).all()}
    # A reference must not point at a row being deleted in the same save.
    return ids - deleted.get(_model_table_key(ref.model), set())


def _name_map(db: Session, ref_key: str) -> dict[str, int]:
    ref = REFS[ref_key]
    return {getattr(o, ref.value_attr): o.id for o in db.scalars(select(ref.model))}


def _model_table_key(model: type) -> str:
    return next((k for k, s in TABLES.items() if s.model is model), "")


_UNSET = object()


def _column_info(model: type) -> dict[str, tuple[bool, object]]:
    """attr -> (nullable, scalar_default). Lets a blank cell fall back to the
    DB column's own default instead of writing NULL into a NOT NULL column."""
    info: dict[str, tuple[bool, object]] = {}
    for col in model.__table__.columns:
        scalar = _UNSET
        d = col.default
        if d is not None and getattr(d, "is_scalar", False):
            scalar = d.arg
        info[col.key] = (col.nullable, scalar)
    return info


def _coerce_scalar(raw: str, col: Column):
    raw = (raw or "").strip()
    if raw == "":
        return None
    if col.type == "int":
        return int(raw)
    if col.type == "bool":
        return _bool(raw)
    if col.type == "date":
        return _parse_date(raw)
    if col.type == "datetime":
        return _parse_datetime(raw)
    return raw


def validate(
    db: Session, spec: TableSpec, rows: list[dict], deletes: list[int]
) -> tuple[list[Coerced], list[dict], dict]:
    """Pure — never writes. Returns (coerced rows, errors, summary)."""
    errors: list[dict] = []
    existing_ids = {r[0] for r in db.execute(select(spec.model.id)).all()}
    deleted_by_table = {spec.key: set(deletes)}
    col_info = _column_info(spec.model)

    # pre-build reference lookups once
    id_ref_sets: dict[str, set[int]] = {}
    name_ref_maps: dict[str, dict[str, int]] = {}
    for col in spec.columns:
        if not col.ref:
            continue
        ref = REFS[col.ref]
        if ref.value_attr == "id":
            id_ref_sets.setdefault(col.ref, _valid_ref_ids(db, col.ref, deleted_by_table))
        else:
            name_ref_maps.setdefault(col.ref, _name_map(db, col.ref))

    coerced: list[Coerced] = []
    adds = updates = 0
    for i, row in enumerate(rows):
        raw_id = str(row.get("id", "")).strip()
        row_id: int | None = int(raw_id) if raw_id else None
        if row_id is not None and row_id not in existing_ids:
            _err(errors, i, "id", f"row id {row_id} does not exist (leave id blank to add a new row)")
            continue
        if row_id is not None and spec.append_only:
            # existing ledger rows are immutable — skip silently, no update
            continue

        fields: dict = {}
        row_ok = True
        for col in spec.columns:
            if not col.editable:
                continue
            try:
                value = _coerce_scalar(row.get(col.name, ""), col)
            except (ValueError, TypeError):
                _err(errors, i, col.name, f"'{row.get(col.name)}' is not a valid {col.type}")
                row_ok = False
                continue

            if value is None and col.required:
                _err(errors, i, col.name, f"{col.name} is required")
                row_ok = False
                continue
            if col.choices and value is not None and value not in col.choices:
                _err(errors, i, col.name, f"{value!r} must be one of: {', '.join(col.choices)}")
                row_ok = False
                continue

            if col.ref and value is not None:
                ref = REFS[col.ref]
                if ref.value_attr == "id":
                    if value not in id_ref_sets[col.ref]:
                        _err(errors, i, col.name, f"{col.name}={value} does not resolve")
                        row_ok = False
                        continue
                    fields[col.model_attr] = value
                else:
                    mapped = name_ref_maps[col.ref].get(value)
                    if mapped is None:
                        _err(errors, i, col.name, f"unknown {col.ref}: {value!r}")
                        row_ok = False
                        continue
                    fields[col.model_attr] = mapped
            elif value is None:
                # a blank cell: keep NULL if the column allows it, otherwise fall
                # back to the column's own default rather than writing NULL. If a
                # NOT NULL column has no scalar default, leave it unset (insert
                # uses the model default; update keeps the existing value).
                nullable, scalar = col_info.get(col.model_attr, (True, _UNSET))
                if nullable:
                    fields[col.model_attr] = None
                elif scalar is not _UNSET:
                    fields[col.model_attr] = scalar() if callable(scalar) else scalar
                # else: skip -> not written
            else:
                fields[col.model_attr] = value

        if not row_ok:
            continue
        coerced.append(Coerced(index=i, id=row_id, fields=fields))
        if row_id is None:
            adds += 1
        else:
            updates += 1

    # deletes
    valid_deletes = 0
    for did in deletes:
        if spec.append_only or spec.delete == "none":
            _err(errors, None, None, f"{spec.label} rows cannot be deleted (append-only ledger)")
            break
        if did not in existing_ids:
            _err(errors, None, None, f"cannot delete id {did}: it does not exist")
            continue
        guard = spec.delete_guard(db, did) if spec.delete_guard else None
        if guard:
            _err(errors, None, None, guard)
            continue
        valid_deletes += 1

    summary = {"adds": adds, "updates": updates, "deletes": valid_deletes}
    return coerced, errors, summary


# --- apply ------------------------------------------------------------------
def apply(db: Session, spec: TableSpec, rows: list[dict], deletes: list[int], actor_id: int) -> dict:
    """Re-validate, then upsert + delete in one transaction. Aborts wholesale
    on any error (nothing written)."""
    coerced, errors, _ = validate(db, spec, rows, deletes)
    if errors:
        return {"ok": False, "errors": errors, "summary": {"adds": 0, "updates": 0, "deletes": 0}}

    adds = updates = removed = 0
    for c in coerced:
        if c.id is None:
            db.add(spec.model(**_insert_defaults(spec, c.fields)))
            adds += 1
        else:
            obj = db.get(spec.model, c.id)
            changed = False
            for attr, value in c.fields.items():
                if getattr(obj, attr) != value:
                    setattr(obj, attr, value)
                    changed = True
            if changed:  # untouched rows are a silent no-op — no churn, honest counts
                updates += 1

    for did in deletes:
        obj = db.get(spec.model, did)
        if obj is None:
            continue
        if spec.delete == "deactivate":
            obj.is_active = False
        elif spec.delete == "soft":
            obj.deleted_at = datetime.now(timezone.utc)
        else:  # hard
            db.delete(obj)
        removed += 1

    summary = {"adds": adds, "updates": updates, "deletes": removed}
    if adds or updates or removed:
        sync.mark_dirty(db, spec.key)
        audit.log(db, actor_id, "bulk", "apply", spec.key, None, summary)
    db.commit()
    return {"ok": True, "errors": [], "summary": summary}


def _insert_defaults(spec: TableSpec, fields: dict) -> dict:
    """Fill columns a new row needs but the grid never edits."""
    out = dict(fields)
    if spec.model is User:
        # bulk-created accounts get an unguessable password; they sign in only
        # after being explicitly linked (see auth.router), matching rebuild.
        out.setdefault("hashed_password", hash_password(secrets.token_urlsafe(24)))
    return out


# --- delete guards ----------------------------------------------------------
def _guard_position(db: Session, pid: int) -> str | None:
    if db.scalar(select(func.count()).select_from(Position).where(Position.parent_id == pid)):
        return f"cannot delete position {pid}: it still has child positions — reparent them first"
    if db.scalar(select(func.count()).select_from(PositionOccupant).where(PositionOccupant.position_id == pid)):
        return f"cannot delete position {pid}: it still has occupants — vacate the seat on the Org page first"
    return None


def _guard_event_kind(db: Session, kid: int) -> str | None:
    if db.scalar(select(func.count()).select_from(Event).where(Event.kind_id == kid)):
        return f"cannot delete event kind {kid}: events still use it"
    return None


def _guard_event(db: Session, eid: int) -> str | None:
    if db.scalar(select(func.count()).select_from(EventCategory).where(EventCategory.event_id == eid)):
        return f"cannot delete event {eid}: delete its categories first"
    return None


def _guard_event_category(db: Session, cid: int) -> str | None:
    if db.scalar(select(func.count()).select_from(EventTeam).where(EventTeam.category_id == cid)):
        return f"cannot delete category {cid}: delete its teams first"
    return None


TABLES["positions"].delete_guard = _guard_position
TABLES["event_kinds"].delete_guard = _guard_event_kind
TABLES["events"].delete_guard = _guard_event
TABLES["event_categories"].delete_guard = _guard_event_category


# --- Excel round-trip -------------------------------------------------------
# The whole workflow: download a real .xlsx of a table, edit it freely in Excel
# or Google Sheets, upload it back. We diff the file against the DB and apply
# adds/updates (and, opt-in, delete rows the user removed from the sheet).
def workbook_bytes(db: Session, spec: TableSpec) -> bytes:
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header = [c.name for c in spec.columns]
    _, rows = _export_for(db, spec)

    wb = Workbook()
    ws = wb.active
    ws.title = spec.key[:31]  # Excel caps sheet names at 31 chars
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    for i, name in enumerate(header, start=1):
        longest = max([len(name)] + [len(str(r[i - 1])) for r in rows]) if rows else len(name)
        ws.column_dimensions[get_column_letter(i)].width = min(longest + 2, 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_upload(content: bytes, filename: str, spec: TableSpec) -> list[dict]:
    """Read an uploaded .xlsx/.csv into row dicts keyed by this table's columns.
    The file must keep every column from the downloaded template (so an update
    never silently blanks a column that was dropped from the sheet)."""
    import csv
    import io

    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[spec.key] if spec.key in wb.sheetnames else wb[wb.sheetnames[0]]
        grid = [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    elif lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        grid = [[c.strip() for c in row] for row in csv.reader(io.StringIO(text))]
    else:
        raise ValueError("Unsupported file — upload a .xlsx or .csv file.")

    # find the header row (the first one containing 'id')
    header_idx = next((i for i, row in enumerate(grid[:25]) if "id" in [c.strip() for c in row]), None)
    if header_idx is None:
        raise ValueError("Could not find the header row — keep the first row of column names.")
    header = [c.strip() for c in grid[header_idx]]
    expected = [c.name for c in spec.columns]
    missing = [c for c in expected if c not in header]
    if missing:
        raise ValueError(f"The sheet is missing column(s): {', '.join(missing)}. Keep every column.")
    idx = {name: header.index(name) for name in expected}

    rows: list[dict] = []
    for raw in grid[header_idx + 1 :]:
        if not any(c.strip() for c in raw):
            continue  # skip fully blank lines
        rows.append({name: (raw[idx[name]] if idx[name] < len(raw) else "") for name in expected})
    return rows


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


# --- Google Sheets escape hatch ---------------------------------------------
# Same round-trip as the Excel one, but the file *is* a live Google Sheet: push
# the current DB into a per-table tab, edit it in Google's UI, then pull it back
# through the exact same validate/apply path. Reuses the service-account client
# the Sheets mirror already uses (app/core/gsheets.py).
BULK_SHEET_BANNER = (
    "[ LIVE — edits here save to the portal automatically. Keep the header row and the "
    "id column; leave id blank to add a row (its id fills in once it saves). ]"
)


def sheet_status() -> dict:
    """Whether the server can talk to Google Sheets at all, plus the target
    spreadsheet's URL. Drives the enabled/disabled state of the Sheets buttons."""
    from app.core import gsheets
    from app.core.config import settings

    sid = settings.google_sheets_spreadsheet_id
    return {
        "configured": gsheets.credentials_available() and bool(sid),
        "spreadsheet_id": sid,
        "url": f"https://docs.google.com/spreadsheets/d/{sid}/edit" if sid else "",
    }


def push_to_sheet(db: Session, spec: TableSpec) -> dict:
    """Overwrite this table's tab with the current DB rows and return a deep
    link to that tab. Raises RuntimeError if Sheets isn't configured."""
    import gspread

    from app.core import gsheets
    from app.core.config import settings

    sid = settings.google_sheets_spreadsheet_id
    if not (gsheets.credentials_available() and sid):
        raise RuntimeError("Google Sheets isn't configured on the server.")

    header = [c.name for c in spec.columns]
    _, rows = _export_for(db, spec)
    ss = gsheets.open_spreadsheet(sid)
    try:
        ws = ss.worksheet(spec.key)
    except gspread.WorksheetNotFound:
        ws = ss.add_worksheet(title=spec.key, rows=max(100, len(rows) + 2), cols=max(len(header), 1))
    grid = [[BULK_SHEET_BANNER] + [""] * (len(header) - 1), header, *rows]
    ws.clear()
    ws.update(grid, "A1")
    try:
        ws.freeze(rows=2)  # keep the banner + header visible while scrolling
    except Exception:  # noqa: BLE001 — freezing is cosmetic, never fatal
        pass
    return {
        "url": f"https://docs.google.com/spreadsheets/d/{sid}/edit#gid={ws.id}",
        "tab": spec.key,
        "rows": len(rows),
    }


def webhook_apply(db: Session, spec: TableSpec, row: dict, actor_id: int | None) -> dict:
    """Apply ONE row edited live in Google Sheets (called by the spreadsheet's
    bound Apps Script on every edit). Same validation as the grid/upload, but
    single-row and it returns the row's id so the script can write it back into
    a freshly-inserted row (blank-id) — which is what stops re-edits of a new
    row from inserting duplicates."""
    coerced, errors, _ = validate(db, spec, [row], [])
    if errors:
        return {"ok": False, "errors": errors, "id": None}
    if not coerced:
        # nothing to write (e.g. an append-only ledger's existing row is immutable)
        raw_id = str(row.get("id", "")).strip()
        return {"ok": True, "errors": [], "id": int(raw_id) if raw_id else None}

    c = coerced[0]
    if c.id is None:
        obj = spec.model(**_insert_defaults(spec, c.fields))
        db.add(obj)
        db.flush()
        new_id, action = obj.id, "insert"
    else:
        obj = db.get(spec.model, c.id)
        for attr, value in c.fields.items():
            if getattr(obj, attr) != value:
                setattr(obj, attr, value)
        new_id, action = c.id, "update"

    sync.mark_dirty(db, spec.key)
    audit.log(db, actor_id, "bulk", f"sheet_{action}", spec.key, new_id, {"via": "sheet-webhook"})
    db.commit()
    return {"ok": True, "errors": [], "id": new_id}


def pull_from_sheet(
    db: Session, spec: TableSpec, do_apply: bool, delete_missing: bool, actor_id: int
) -> dict:
    """Read this table's tab back and run it through the same diff/apply path as
    an uploaded file. `do_apply=False` previews; `True` commits."""
    from app.core import gsheets
    from app.core.config import settings

    sid = settings.google_sheets_spreadsheet_id
    if not (gsheets.credentials_available() and sid):
        raise RuntimeError("Google Sheets isn't configured on the server.")

    header, sheet_rows = gsheets.read_worksheet(sid, spec.key)
    expected = [c.name for c in spec.columns]
    missing = [c for c in expected if c not in header]
    if missing:
        raise ValueError(
            f"The '{spec.key}' tab is missing column(s): {', '.join(missing)}. "
            "Keep every column — re-open it from the portal to reset the tab."
        )
    rows = [{name: r.get(name, "") for name in expected} for r in sheet_rows]

    plan = plan_upload(db, spec, rows, delete_missing)
    if not do_apply or not plan["ok"]:
        return {"applied": False, **plan}
    deletes = plan["missing_ids"] if (delete_missing and plan["can_delete"]) else []
    result = apply(db, spec, rows, deletes, actor_id)
    return {"applied": result["ok"], **result}


def plan_upload(db: Session, spec: TableSpec, rows: list[dict], delete_missing: bool) -> dict:
    """Diff the uploaded rows against the DB without writing. Returns the counts
    the confirm dialog shows, the ids that would be deleted, and any errors."""
    coerced, errors, _ = validate(db, spec, rows, [])
    existing_ids = {r[0] for r in db.execute(select(spec.model.id)).all()}
    uploaded_ids = {c.id for c in coerced if c.id is not None}

    new = changed = unchanged = 0
    for c in coerced:
        if c.id is None:
            new += 1
            continue
        obj = db.get(spec.model, c.id)
        if obj is not None and any(getattr(obj, a) != v for a, v in c.fields.items()):
            changed += 1
        else:
            unchanged += 1

    # rows in the DB but absent from the uploaded sheet — candidate deletions
    can_delete = not (spec.append_only or spec.delete == "none")
    missing_ids = sorted(existing_ids - uploaded_ids) if can_delete else []
    return {
        "ok": not errors,
        "errors": errors,
        "new": new,
        "changed": changed,
        "unchanged": unchanged,
        "missing_ids": missing_ids,
        "can_delete": can_delete,
        "will_delete": len(missing_ids) if delete_missing else 0,
    }
